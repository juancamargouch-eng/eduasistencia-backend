from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, joinedload
from fastapi.responses import StreamingResponse
from ...config.database import get_db
from ...models.attendance import AttendanceLog
from ...models.student import Student
import pandas as pd
import io
from datetime import date

from ...api import deps
from ...models.user import User

router = APIRouter()

@router.get("/attendance/export")
def export_attendance(
    from_date: date,
    to_date: date,
    grade: str = None,
    section: str = None,
    search: str = None,
    status_filter: str = None,
    schedule_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    try:
        def clean_filter(value):
            if value is None: return None
            s = str(value).strip()
            if s.lower() in ["", "undefined", "null", "none", "todos", "todas"]:
                return None
            return s

        grade = clean_filter(grade)
        section = clean_filter(section)
        search = clean_filter(search)
        status_filter = clean_filter(status_filter)
        
        print(f"DEBUG: EXPORT - From:{from_date} To:{to_date} G:{grade} S:{section} Search:{search} Status:{status_filter} Sch:{schedule_id}")
        
        from sqlalchemy import Date, cast, or_
        import re

        # 1. Get students matching filters
        student_query = db.query(Student).options(joinedload(Student.schedule))
        if grade: student_query = student_query.filter(Student.grade == grade)
        if section: student_query = student_query.filter(Student.section == section)
        if schedule_id: student_query = student_query.filter(Student.schedule_id == schedule_id)
        if search:
            pattern = f"%{search}%"
            student_query = student_query.filter(or_(Student.full_name.ilike(pattern), Student.dni.ilike(pattern)))
        
        students = student_query.all()
        if not students:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame(columns=["Mensaje"]).to_excel(writer, index=False, sheet_name='Sin Resultados')
            output.seek(0)
            return StreamingResponse(output, headers={'Content-Disposition': 'attachment; filename="reporte_vacio.xlsx"'}, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        student_ids = [s.id for s in students]

        # 2. Get all logs in range for these students (we don't filter status yet to avoid gaps in matrix)
        log_query = db.query(AttendanceLog).filter(AttendanceLog.student_id.in_(student_ids))
        log_query = log_query.filter(cast(AttendanceLog.timestamp, Date) >= from_date)
        log_query = log_query.filter(cast(AttendanceLog.timestamp, Date) <= to_date)
        
        # Sort by timestamp to handle earliest movement in matrix
        logs = log_query.order_by(AttendanceLog.timestamp.asc()).all()
        print(f"DEBUG: Logs found for students: {len(logs)}")

        # 2.5 Get approved justifications for these students
        from ...models.justification import Justification, JustificationStatus
        just_query = db.query(Justification).filter(
            Justification.student_id.in_(student_ids),
            Justification.date >= from_date,
            Justification.date <= to_date,
            Justification.status == JustificationStatus.APPROVED
        )
        justifications = just_query.all()
        just_map = {(j.student_id, j.date): j.reason for j in justifications}
        
        # 3. Create a mapping (student_id, date) -> log
        from collections import defaultdict
        attendance_map = defaultdict(lambda: None)
        for log in logs:
            if not log.timestamp: continue
            
            # Use local date comparison
            log_date = log.timestamp.date()
            key = (log.student_id, log_date)
            
            existing = attendance_map[key]
            # Priority: 1. Successful logs, 2. Earliest logs if same status
            if not existing:
                attendance_map[key] = log
            else:
                if not existing.verification_status and log.verification_status:
                    attendance_map[key] = log
                elif existing.verification_status == log.verification_status:
                    # If both have same success status, we keep the earliest one for Entry
                    pass # Already have the earliest due to ORDER BY asc

        # 4. Generate Data
        import datetime
        date_range = []
        curr = from_date
        while curr <= to_date:
            date_range.append(curr)
            curr += datetime.timedelta(days=1)

        # 5. Build Final Data
        final_rows = []
        for student in students:
            for d in date_range:
                log = attendance_map[(student.id, d)]
                status = "Falta"
                hora = "--:--:--"
                tipo = "Entrada"
                
                if log:
                    tipo = "Entrada" if log.event_type == "ENTRY" else "Salida"
                    # Robust time extract
                    try:
                        hora = log.timestamp.strftime("%H:%M:%S")
                    except:
                        hora = "S/H"
                        
                    if log.verification_status:
                        status = "Tardanza" if log.status == "LATE" else "Presente"
                    else:
                        status = f"Fallido ({log.failure_reason or 'Error'})"
                else:
                    # CHECK IF JUSTIFIED
                    justification_reason = just_map.get((student.id, d))
                    if justification_reason:
                        status = f"Justificada"
                        hora = "JUSTIFIC."

                # APPLY STATUS FILTER AFTER MATRIX GENERATION
                # This ensures that "Falta" is also searchable/filterable correctly
                if status_filter:
                    current_status_upper = status.upper()
                    filter_upper = status_filter.upper()
                    
                    # Logic: If filtering for 'FALTA', show only faltas
                    # If filtering for 'PRESENTE', show only 'PRESENTE'
                    # etc.
                    match = False
                    if filter_upper == "FALTA" and current_status_upper == "FALTA": match = True
                    elif filter_upper == "TARDANZA" and current_status_upper == "TARDANZA": match = True
                    elif filter_upper == "PRESENTE" and current_status_upper == "PRESENTE": match = True
                    # Partial match for failed attempts
                    elif filter_upper == "FALLIDO" and "FALLIDO" in current_status_upper: match = True
                    elif filter_upper == "JUSTIFICADA" and "JUSTIFICADA" in current_status_upper: match = True
                    
                    if not match: continue

                final_rows.append({
                    "DNI": student.dni,
                    "Nombre": student.full_name,
                    "Grado": student.grade,
                    "Sección": student.section,
                    "Fecha": d.strftime("%Y-%m-%d"),
                    "Hora": hora,
                    "Turno": student.schedule.name if student.schedule else "N/A",
                    "Estado": status,
                    "Tipo": tipo
                })

        df = pd.DataFrame(final_rows)
        if df.empty:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame(columns=["Mensaje"]).to_excel(writer, index=False, sheet_name='Sin Resultados')
                ws = writer.sheets['Sin Resultados']
                ws['A1'] = "No se encontraron registros de asistencia con los filtros seleccionados."
                ws.column_dimensions['A'].width = 60
            output.seek(0)
            return StreamingResponse(output, headers={'Content-Disposition': 'attachment; filename="reporte_vacio.xlsx"'}, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        def safe_sheet_name(name):
            if not name: return "General"
            clean = re.sub(r'[\\/*\[\]:?]', '-', str(name))
            return clean[:30] if clean else "General"

        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                grades = sorted(df['Grado'].unique(), key=lambda x: str(x))
                for g in grades:
                    try:
                        grade_label = str(g) if g else "General"
                        df_grade = df[df['Grado'] == g].sort_values(by=['Fecha', 'Sección', 'Nombre'])
                        sheet_name = safe_sheet_name(grade_label)
                        
                        # Write data starting from row 2
                        df_grade.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
                        ws = writer.sheets[sheet_name]
                        
                        # 1. Title Row (Merged)
                        try:
                            ws.merge_cells('A1:I1')
                            title_cell = ws['A1']
                            title_cell.value = f"VerifID - Registro de Control de Asistencia | {grade_label}"
                            title_cell.font = Font(color="FFFFFF", bold=True, size=14)
                            title_cell.fill = PatternFill(start_color="FFB2041A", end_color="FFB2041A", fill_type="solid")
                            title_cell.alignment = Alignment(horizontal="center", vertical="center")
                            ws.row_dimensions[1].height = 30
                        except Exception as e_title:
                            print(f"Error drawing title: {e_title}")

                        # 2. Header Style (Row 2)
                        try:
                            header_fill = PatternFill(start_color="FFE10521", end_color="FFE10521", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            for cell in ws[2]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        except Exception as e_header:
                            print(f"Error drawing header: {e_header}")

                        # 3. Data Styling (Row 3 onwards)
                        try:
                            for row in ws.iter_rows(min_row=3):
                                for cell in row:
                                    # Alignment for specific columns
                                    if cell.column_letter in ['E', 'F', 'G', 'H']: 
                                        cell.alignment = Alignment(horizontal="center")
                                    
                                    # Conditional Formatting for Status (Column H)
                                    if cell.column_letter == 'H':
                                        val = str(cell.value or "").upper()
                                        if any(x in val for x in ["TARDANZA", "FALTA", "FALLIDO"]):
                                            cell.font = Font(color="FFE10521", bold=True)
                                        elif "PRESENTE" in val:
                                            cell.font = Font(color="FF008000", bold=True)
                                        elif "JUSTIFICADA" in val:
                                            cell.font = Font(color="FF0000FF", bold=True) # Blue for Justified
                        except Exception as e_data:
                            print(f"Error styling data: {e_data}")

                        # 4. Auto-width (Start from Row 2 to avoid merged cell A1)
                        try:
                            for col_idx in range(1, 10): # A to I
                                max_len = 0
                                for row_idx in range(2, ws.max_row + 1):
                                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                                    if cell_val:
                                        max_len = max(max_len, len(str(cell_val)))
                                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
                        except Exception as e_width:
                            print(f"Error setting width: {e_width}")
                            
                    except Exception as e_sheet:
                        print(f"Error processing sheet {g}: {e_sheet}")
                        
        except Exception as e_excel:
            import traceback
            traceback.print_exc()
            print(f"Critical Excel error: {e_excel}")
            raise HTTPException(status_code=500, detail=f"Error al generar archivo Excel: {str(e_excel)}")

        output.seek(0)
        return StreamingResponse(
            output, 
            headers={'Content-Disposition': f'attachment; filename="VerifID_Reporte_{from_date}.xlsx"'}, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
